from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, HttpResponseBadRequest
import csv
import io
import openpyxl
from .models import Station, Bay, Relay
from django.db.models import Q

def is_admin(user):
    return user.is_superuser or user.groups.filter(name='ADMIN').exists()

@login_required
def station_list(request):
    q = request.GET.get('q', '').strip()
    
    if q:
        stations = Station.objects.prefetch_related('bays__relays__settingsheet_set').filter(
            Q(station_code__icontains=q) | 
            Q(station_name__icontains=q) |
            Q(bays__bay_code__icontains=q) |
            Q(bays__bay_name__icontains=q) |
            Q(bays__relays__relay_code__icontains=q) |
            Q(bays__relays__relay_name__icontains=q)
        ).distinct()
        is_search = True
    else:
        stations = Station.objects.all()
        is_search = False

    return render(request, 'stations/station_list.html', {
        'stations': stations,
        'q': q,
        'is_search': is_search
    })

@login_required
@user_passes_test(is_admin)
def station_create(request):
    if request.method == 'POST':
        station_code = request.POST.get('station_code')
        station_name = request.POST.get('station_name')
        location = request.POST.get('location', '')
        if station_code and station_name:
            Station.objects.create(station_code=station_code, station_name=station_name, location=location)
            messages.success(request, f"Đã tạo trạm {station_name}.")
        else:
            messages.error(request, "Vui lòng điền đủ Mã Trạm và Tên Trạm.")
    return redirect('station_list')

@login_required
@user_passes_test(is_admin)
def bay_create(request, station_id):
    if request.method == 'POST':
        station = get_object_or_404(Station, pk=station_id)
        bay_code = request.POST.get('bay_code')
        bay_name = request.POST.get('bay_name', '')
        if bay_code:
            Bay.objects.create(station=station, bay_code=bay_code, bay_name=bay_name)
            messages.success(request, f"Đã thêm ngăn lộ {bay_code} vào trạm {station.station_code}.")
        else:
            messages.error(request, "Vui lòng điền Mã Ngăn lộ.")
    return redirect('station_list')

@login_required
@user_passes_test(is_admin)
def relay_create(request, bay_id):
    if request.method == 'POST':
        bay = get_object_or_404(Bay, pk=bay_id)
        relay_code = request.POST.get('relay_code')
        relay_name = request.POST.get('relay_name', '')
        manufacturer = request.POST.get('manufacturer', '')
        if relay_code:
            Relay.objects.create(bay=bay, relay_code=relay_code, relay_name=relay_name, manufacturer=manufacturer)
            messages.success(request, f"Đã thêm Rơ-le {relay_code} vào ngăn lộ {bay.bay_code}.")
        else:
            messages.error(request, "Vui lòng điền Mã Rơ-le.")
    return redirect('station_list')


@login_required
@user_passes_test(is_admin)
def relay_bulk_create(request, bay_id):
    if request.method == 'POST':
        bay = get_object_or_404(Bay, pk=bay_id)
        created_count = 0
        
        # Xử lý File Upload (CSV)
        if 'csv_file' in request.FILES:
            csv_file = request.FILES['csv_file']
            if not csv_file.name.endswith('.csv'):
                messages.error(request, "Vui lòng tải lên file định dạng CSV.")
                return redirect('station_list')
                
            try:
                decoded_file = csv_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string, delimiter=',')
                relays_to_create = []
                for idx, row in enumerate(reader):
                    if idx == 0: continue
                    if len(row) >= 1 and row[0].strip():
                        r_code = row[0].strip()
                        r_name = row[1].strip() if len(row) > 1 else ''
                        r_manuf = row[2].strip() if len(row) > 2 else ''
                        relays_to_create.append(Relay(bay=bay, relay_code=r_code, relay_name=r_name, manufacturer=r_manuf))
                
                Relay.objects.bulk_create(relays_to_create, ignore_conflicts=True)
                created_count = len(relays_to_create)
            except Exception as e:
                messages.error(request, f"Lỗi đọc file CSV: {str(e)}")
                return redirect('station_list')

        # Xử lý Paste Text (Raw Text)
        elif 'raw_text' in request.POST:
            raw_text = request.POST.get('raw_text', '')
            relays_to_create = []
            lines = raw_text.strip().split('\n')
            for line in lines:
                if not line.strip(): continue
                cols = line.split('\t')
                r_code = cols[0].strip()
                r_name = cols[1].strip() if len(cols) > 1 else ''
                r_manuf = cols[2].strip() if len(cols) > 2 else ''
                if r_code:
                    relays_to_create.append(Relay(bay=bay, relay_code=r_code, relay_name=r_name, manufacturer=r_manuf))
            
            Relay.objects.bulk_create(relays_to_create, ignore_conflicts=True)
            created_count = len(relays_to_create)
            
        messages.success(request, f"Đã nhập thành công {created_count} Rơ-le vào ngăn lộ {bay.bay_code}.")
    return redirect('station_list')

@login_required
@user_passes_test(is_admin)
def global_bulk_create(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            messages.error(request, "Vui lòng tải lên file định dạng Excel (.xlsx hoặc .xls).")
            return redirect('station_list')
            
        stations_cache = {}
        bays_cache = {}
        relays_to_create = []
        created_count = 0
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0: continue # Skip header
                if not row or len(row) < 5 or not row[0]: continue
                
                s_code = str(row[0]).strip()
                s_name = str(row[1]).strip() if row[1] else s_code
                b_code = str(row[2]).strip()
                b_name = str(row[3]).strip() if row[3] else b_code
                r_code = str(row[4]).strip()
                r_name = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                r_manuf = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                
                # Get or Create Station
                if s_code not in stations_cache:
                    station, _ = Station.objects.get_or_create(station_code=s_code, defaults={'station_name': s_name})
                    stations_cache[s_code] = station
                station = stations_cache[s_code]
                
                # Get or Create Bay
                cache_key = f"{s_code}_{b_code}"
                if cache_key not in bays_cache:
                    bay, _ = Bay.objects.get_or_create(station=station, bay_code=b_code, defaults={'bay_name': b_name})
                    bays_cache[cache_key] = bay
                bay = bays_cache[cache_key]
                
                # Add to relay list
                if r_code:
                    relays_to_create.append(Relay(bay=bay, relay_code=r_code, relay_name=r_name, manufacturer=r_manuf))
            
            if relays_to_create:
                Relay.objects.bulk_create(relays_to_create, ignore_conflicts=True)
                created_count = len(relays_to_create)
            
            messages.success(request, f"Đã nhập toàn hệ thống từ Excel thành công! Tạo mới {len(stations_cache)} Trạm, {len(bays_cache)} Ngăn lộ và {created_count} Rơ-le.")
        except Exception as e:
            messages.error(request, f"Có lỗi xảy ra khi đọc file Excel: {str(e)}")
            
    return redirect('station_list')


@login_required
def relay_status_dashboard(request):
    """Bảng điều khiển theo dõi tiến độ cấu hình rơ-le"""
    from django.core.paginator import Paginator

    # Tất cả rơ-le
    total_relays = Relay.objects.count()
    
    # Rơ-le đã có phiếu
    configured_relays_qs = Relay.objects.filter(settingsheet__isnull=False).distinct().select_related('bay', 'bay__station').order_by('relay_code')
    configured_count = configured_relays_qs.count()
    
    # Rơ-le chưa có phiếu
    unconfigured_relays_qs = Relay.objects.filter(settingsheet__isnull=True).select_related('bay', 'bay__station').order_by('relay_code')
    unconfigured_count = total_relays - configured_count
    
    # Phân trang (40 items/trang)
    page_unc = request.GET.get('page_unc', 1)
    paginator_unc = Paginator(unconfigured_relays_qs, 40)
    unconfigured_relays = paginator_unc.get_page(page_unc)

    page_con = request.GET.get('page_con', 1)
    paginator_con = Paginator(configured_relays_qs, 40)
    configured_relays = paginator_con.get_page(page_con)

    # Active Tab từ Request
    active_tab = request.GET.get('tab', 'unconfigured')

    # Phần trăm
    progress_percent = int((configured_count / total_relays * 100) if total_relays > 0 else 0)
    
    context = {
        'total_relays': total_relays,
        'configured_count': configured_count,
        'unconfigured_count': unconfigured_count,
        'progress_percent': progress_percent,
        'configured_relays': configured_relays,
        'unconfigured_relays': unconfigured_relays,
        'active_tab': active_tab
    }
    
    return render(request, 'stations/relay_status.html', context)

@login_required
def get_bays_htmx(request, station_id):
    station = get_object_or_404(Station, pk=station_id)
    bays = station.bays.all()
    return render(request, 'stations/partials/bay_list.html', {'bays': bays, 'station': station})

@login_required
def get_relays_htmx(request, bay_id):
    bay = get_object_or_404(Bay, pk=bay_id)
    relays = bay.relays.prefetch_related('settingsheet_set').all()
    return render(request, 'stations/partials/relay_list.html', {'relays': relays, 'bay': bay})

@login_required
def search_suggestions(request):
    q = request.GET.get('q', '').strip()
    suggestions = []
    if q and len(q) >= 1:
        st = Station.objects.filter(Q(station_code__icontains=q) | Q(station_name__icontains=q))[:3]
        for s in st: suggestions.append(s.station_code)
        
        by = Bay.objects.filter(Q(bay_code__icontains=q) | Q(bay_name__icontains=q))[:3]
        for b in by: suggestions.append(b.bay_code)
        
        rl = Relay.objects.filter(Q(relay_code__icontains=q) | Q(relay_name__icontains=q))[:5]
        for r in rl: suggestions.append(r.relay_code)
    return render(request, 'stations/partials/suggestions.html', {'suggestions': suggestions})

def relay_create(request, bay_id):
    if request.method == 'POST':
        bay = get_object_or_404(Bay, pk=bay_id)
        relay_code = request.POST.get('relay_code')
        relay_name = request.POST.get('relay_name', '')
        manufacturer = request.POST.get('manufacturer', '')
        if relay_code:
            Relay.objects.create(bay=bay, relay_code=relay_code, relay_name=relay_name, manufacturer=manufacturer)
            messages.success(request, f"Đã thêm Rơ-le {relay_code} vào ngăn lộ {bay.bay_code}.")
        else:
            messages.error(request, "Vui lòng điền Mã Rơ-le.")
    return redirect('station_list')


@login_required
@user_passes_test(is_admin)
def relay_bulk_create(request, bay_id):
    if request.method == 'POST':
        bay = get_object_or_404(Bay, pk=bay_id)
        created_count = 0
        
        # Xử lý File Upload (CSV)
        if 'csv_file' in request.FILES:
            csv_file = request.FILES['csv_file']
            if not csv_file.name.endswith('.csv'):
                messages.error(request, "Vui lòng tải lên file định dạng CSV.")
                return redirect('station_list')
                
            try:
                decoded_file = csv_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string, delimiter=',')
                relays_to_create = []
                for idx, row in enumerate(reader):
                    if idx == 0: continue
                    if len(row) >= 1 and row[0].strip():
                        r_code = row[0].strip()
                        r_name = row[1].strip() if len(row) > 1 else ''
                        r_manuf = row[2].strip() if len(row) > 2 else ''
                        relays_to_create.append(Relay(bay=bay, relay_code=r_code, relay_name=r_name, manufacturer=r_manuf))
                
                Relay.objects.bulk_create(relays_to_create, ignore_conflicts=True)
                created_count = len(relays_to_create)
            except Exception as e:
                messages.error(request, f"Lỗi đọc file CSV: {str(e)}")
                return redirect('station_list')

        # Xử lý Paste Text (Raw Text)
        elif 'raw_text' in request.POST:
            raw_text = request.POST.get('raw_text', '')
            relays_to_create = []
            lines = raw_text.strip().split('\n')
            for line in lines:
                if not line.strip(): continue
                cols = line.split('\t')
                r_code = cols[0].strip()
                r_name = cols[1].strip() if len(cols) > 1 else ''
                r_manuf = cols[2].strip() if len(cols) > 2 else ''
                if r_code:
                    relays_to_create.append(Relay(bay=bay, relay_code=r_code, relay_name=r_name, manufacturer=r_manuf))
            
            Relay.objects.bulk_create(relays_to_create, ignore_conflicts=True)
            created_count = len(relays_to_create)
            
        messages.success(request, f"Đã nhập thành công {created_count} Rơ-le vào ngăn lộ {bay.bay_code}.")
    return redirect('station_list')

@login_required
@user_passes_test(is_admin)
def global_bulk_create(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            messages.error(request, "Vui lòng tải lên file định dạng Excel (.xlsx hoặc .xls).")
            return redirect('station_list')
            
        stations_cache = {}
        bays_cache = {}
        relays_to_create = []
        created_count = 0
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0: continue # Skip header
                if not row or len(row) < 5 or not row[0]: continue
                
                s_code = str(row[0]).strip()
                s_name = str(row[1]).strip() if row[1] else s_code
                b_code = str(row[2]).strip()
                b_name = str(row[3]).strip() if row[3] else b_code
                r_code = str(row[4]).strip()
                r_name = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                r_manuf = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                
                # Get or Create Station
                if s_code not in stations_cache:
                    station, _ = Station.objects.get_or_create(station_code=s_code, defaults={'station_name': s_name})
                    stations_cache[s_code] = station
                station = stations_cache[s_code]
                
                # Get or Create Bay
                cache_key = f"{s_code}_{b_code}"
                if cache_key not in bays_cache:
                    bay, _ = Bay.objects.get_or_create(station=station, bay_code=b_code, defaults={'bay_name': b_name})
                    bays_cache[cache_key] = bay
                bay = bays_cache[cache_key]
                
                # Add to relay list
                if r_code:
                    relays_to_create.append(Relay(bay=bay, relay_code=r_code, relay_name=r_name, manufacturer=r_manuf))
            
            if relays_to_create:
                Relay.objects.bulk_create(relays_to_create, ignore_conflicts=True)
                created_count = len(relays_to_create)
            
            messages.success(request, f"Đã nhập toàn hệ thống từ Excel thành công! Tạo mới {len(stations_cache)} Trạm, {len(bays_cache)} Ngăn lộ và {created_count} Rơ-le.")
        except Exception as e:
            messages.error(request, f"Có lỗi xảy ra khi đọc file Excel: {str(e)}")
            
    return redirect('station_list')


@login_required
def relay_status_dashboard(request):
    """Bảng điều khiển theo dõi tiến độ cấu hình rơ-le"""
    from django.core.paginator import Paginator

    # Tất cả rơ-le
    total_relays = Relay.objects.count()
    
    # Rơ-le đã có phiếu
    configured_relays_qs = Relay.objects.filter(settingsheet__isnull=False).distinct().select_related('bay', 'bay__station').order_by('relay_code')
    configured_count = configured_relays_qs.count()
    
    # Rơ-le chưa có phiếu
    unconfigured_relays_qs = Relay.objects.filter(settingsheet__isnull=True).select_related('bay', 'bay__station').order_by('relay_code')
    unconfigured_count = total_relays - configured_count
    
    # Phân trang (40 items/trang)
    page_unc = request.GET.get('page_unc', 1)
    paginator_unc = Paginator(unconfigured_relays_qs, 40)
    unconfigured_relays = paginator_unc.get_page(page_unc)

    page_con = request.GET.get('page_con', 1)
    paginator_con = Paginator(configured_relays_qs, 40)
    configured_relays = paginator_con.get_page(page_con)

    # Active Tab từ Request
    active_tab = request.GET.get('tab', 'unconfigured')

    # Phần trăm
    progress_percent = int((configured_count / total_relays * 100) if total_relays > 0 else 0)
    
    context = {
        'total_relays': total_relays,
        'configured_count': configured_count,
        'unconfigured_count': unconfigured_count,
        'progress_percent': progress_percent,
        'configured_relays': configured_relays,
        'unconfigured_relays': unconfigured_relays,
        'active_tab': active_tab
    }
    
    return render(request, 'stations/relay_status.html', context)

@login_required
def get_bays_htmx(request, station_id):
    station = get_object_or_404(Station, pk=station_id)
    bays = station.bays.all()
    return render(request, 'stations/partials/bay_list.html', {'bays': bays, 'station': station})

@login_required
def get_relays_htmx(request, bay_id):
    bay = get_object_or_404(Bay, pk=bay_id)
    relays = bay.relays.prefetch_related('settingsheet_set').all()
    return render(request, 'stations/partials/relay_list.html', {'relays': relays, 'bay': bay})

@login_required
def search_suggestions(request):
    q = request.GET.get('q', '').strip()
    suggestions = []
    if q and len(q) >= 1:
        st = Station.objects.filter(Q(station_code__icontains=q) | Q(station_name__icontains=q))[:3]
        for s in st: suggestions.append(s.station_code)
        
        by = Bay.objects.filter(Q(bay_code__icontains=q) | Q(bay_name__icontains=q))[:3]
        for b in by: suggestions.append(b.bay_code)
        
        rl = Relay.objects.filter(Q(relay_code__icontains=q) | Q(relay_name__icontains=q))[:5]
        for r in rl: suggestions.append(r.relay_code)
    return render(request, 'stations/partials/suggestions.html', {'suggestions': suggestions})

@login_required
@user_passes_test(is_admin)
def run_autocheck_now(request, relay_id):
    from .api_mock import compare_and_log_relay_check
    relay = get_object_or_404(Relay, pk=relay_id)
    
    if request.method == 'POST':
        # Chỉ chạy kiểm tra mới khi POST (nút xanh dương - Sync)
        compare_and_log_relay_check(relay)
        # Refresh relay từ DB để lấy trạng thái mới nhất
        relay.refresh_from_db()
    
    response = render(request, 'stations/partials/_autocheck_results.html', {'relay': relay})
    if request.method == 'POST':
        import json
        response["HX-Trigger"] = json.dumps({f"refreshRow-{relay.id}": ""})
    return response

@login_required
@user_passes_test(is_admin)
def update_relay_schedule(request, relay_id):
    relay = get_object_or_404(Relay, pk=relay_id)
    if request.method == 'POST':
        enabled = request.POST.get('enabled') == 'on'
        schedule_mode = request.POST.get('schedule_mode', 'interval')
        
        relay.auto_check_enabled = enabled
        
        if schedule_mode == 'exact':
            relay.check_interval_unit = 'e'
            relay.check_interval_value = 0
            exact_time_str = request.POST.get('exact_datetime')
            if exact_time_str:
                from django.utils.dateparse import parse_datetime
                from django.utils import timezone
                dt = parse_datetime(exact_time_str)
                if dt:
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt, timezone.get_current_timezone())
                    relay.next_check_at = dt
                else:
                    relay.next_check_at = None
            else:
                relay.next_check_at = None
        else:
            interval_val = request.POST.get('interval_value', 1)
            interval_unit = request.POST.get('interval_unit', 'h')
            try:
                relay.check_interval_value = int(interval_val)
            except ValueError:
                relay.check_interval_value = 1
            relay.check_interval_unit = interval_unit
            
            if enabled:
                relay.next_check_at = relay.calculate_next_check()
            else:
                relay.next_check_at = None
            
        relay.save()
        
        # Clear cache to ensure dashboard statistics update immediately
        from django.core.cache import cache
        cache.delete('autocheck_dashboard_stats')
        
        from django.http import HttpResponse
        import json
        
        # Trigger HTMX events on the client side:
        # 1. notify: to show toast
        # 2. refreshRow-{relay.id}: to update the table row
        triggers = {
            "notify": {"title": "Thành công", "message": "Đã lưu cấu hình rơ-le", "level": "success"},
            f"refreshRow-{relay.id}": ""
        }
        
        response = render(request, 'stations/partials/_schedule_config.html', {'relay': relay})
        response["HX-Trigger"] = json.dumps(triggers)
        return response
        
    return render(request, 'stations/partials/_schedule_config.html', {'relay': relay})

@login_required
@user_passes_test(is_admin)
def autocheck_dashboard(request):
    from django.utils import timezone
    from datetime import timedelta
    from django.core.cache import cache
    from sheets.models import SettingSheet
    
    # Lấy tất cả rơ-le trong database
    qs = Relay.objects.all().select_related('bay', 'bay__station')
    
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    
    if q:
        qs = qs.filter(Q(relay_code__icontains=q) | Q(relay_name__icontains=q) | Q(bay__station__station_name__icontains=q) | Q(bay__station__station_code__icontains=q))
    
    if status == 'enabled':
        qs = qs.filter(auto_check_enabled=True)
    elif status == 'disabled':
        qs = qs.filter(auto_check_enabled=False)
        
    qs = qs.order_by('-auto_check_enabled', 'next_check_at')
    
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    stats = cache.get('autocheck_dashboard_stats')
    if not stats:
        total_relays = Relay.objects.count()
        enabled_count = Relay.objects.filter(auto_check_enabled=True).count()
        disabled_count = total_relays - enabled_count
        
        now = timezone.now()
        one_hour_later = now + timedelta(hours=1)
        soon_count = Relay.objects.filter(
            auto_check_enabled=True,
            next_check_at__lte=one_hour_later
        ).count()
        
        stats = {
            'total_relays': total_relays,
            'enabled_count': enabled_count,
            'disabled_count': disabled_count,
            'soon_count': soon_count
        }
        cache.set('autocheck_dashboard_stats', stats, 300) # Cache for 5 minutes
        
    total_relays = stats['total_relays']
    enabled_count = stats['enabled_count']
    disabled_count = stats['disabled_count']
    soon_count = stats['soon_count']
    
    context = {
        'page_obj': page_obj,
        'q': q,
        'status_filter': status,
        'total_relays': total_relays,
        'enabled_count': enabled_count,
        'disabled_count': disabled_count,
        'soon_count': soon_count
    }
    
    if request.headers.get('HX-Request') == 'true':
        return render(request, 'stations/partials/_autocheck_table.html', context)
        
    return render(request, 'stations/autocheck_dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def workflow_continue(request, relay_id):
    """
    Bỏ qua lỗi và tiếp tục lịch hẹn giờ.
    """
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid method")
        
    relay = get_object_or_404(Relay, pk=relay_id)
    if relay.is_paused_for_correction:
        relay.is_paused_for_correction = False
        relay.paused_schedule_data = None
        
        # Cập nhật lại next_check_at
        from django.utils import timezone
        relay.last_checked_at = timezone.now()
        relay.next_check_at = relay.calculate_next_check(from_time=relay.last_checked_at)
        relay.save(update_fields=['is_paused_for_correction', 'paused_schedule_data', 'last_checked_at', 'next_check_at'])
        
    import json
    triggers = {
        "notify": {"title": "Thành công", "message": "Đã bỏ qua sai lệch, tiếp tục lấy dữ liệu.", "level": "success"},
        f"refreshRow-{relay.id}": "",
        "closeResultModal": ""
    }
    
    # Return empty response with HTMX triggers
    response = HttpResponse("")
    response["HX-Trigger"] = json.dumps(triggers)
    return response

@login_required
@user_passes_test(is_admin)
def workflow_create_ticket(request, relay_id):
    """
    Tạo Phiếu Chỉnh định (SettingSheet) để kích hoạt luồng xử lý lỗi thực tế.
    """
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid method")
        
    relay = get_object_or_404(Relay, pk=relay_id)
    latest_log = relay.auto_checks.first()
    
    # Kích hoạt trạng thái tạm dừng nếu chưa có
    if not relay.is_paused_for_correction:
        relay.is_paused_for_correction = True
        relay.paused_schedule_data = {
            'check_interval_value': relay.check_interval_value,
            'check_interval_unit': relay.check_interval_unit
        }
        relay.save(update_fields=['is_paused_for_correction', 'paused_schedule_data'])
        
    # Tạo SettingSheet (Phiếu Chỉnh định) thực sự
    from sheets.models import SettingSheet
    from django.utils import timezone
    import json
    
    timestamp = timezone.now().strftime('%Y%m%d%H%M')
    sheet_code = f"AUTO-CORRECT-{relay.relay_code}-{timestamp}"
    
    # Convert API raw data to extracted_data format for SettingSheet
    # API raw data is: [{'code': '...', 'name': '...', 'value': '...', 'is_diff': True, 'db_value': '...'}, ...]
    extracted_data = []
    if latest_log and latest_log.api_raw_data:
        for param in latest_log.api_raw_data:
            target_value = param.get('db_value')
            if target_value is None:
                target_value = param.get('value', '')
                
            item_data = {
                'parameter_code': param.get('code'),
                'parameter_name': param.get('name'),
                'value': str(target_value)
            }
            
            # Nếu có sai lệch, lưu lại giá trị sai để hiển thị ở tab So sánh
            if param.get('is_diff'):
                item_data['wrong_value'] = str(param.get('value', ''))
                
            extracted_data.append(item_data)
            
    sheet = SettingSheet.objects.create(
        sheet_code=sheet_code,
        title=f"Phiếu xử lý sai lệch tự động cho {relay.relay_name}",
        status='ROUTED_TO_STATION', # Bỏ qua bước Dispatcher, chuyển thẳng về Trạm
        relay=relay,
        station=relay.bay.station if relay.bay else None,
        created_by=request.user,
        extracted_data=extracted_data
    )
    
    # Xóa cache danh sách phiếu
    from django.core.cache import cache
    import time
    cache.set('sheet_list_version', int(time.time()))
    
    # Trả về thông báo thành công và redirect (hoặc reload)
    triggers = {
        "notify": {"title": "Đã tạo phiếu", "message": f"Đã khởi tạo phiếu {sheet_code} và chuyển cho Phân phối viên.", "level": "success"},
        f"refreshRow-{relay.id}": "",
        "closeResultModal": ""
    }
    
    from django.http import HttpResponse
    response = HttpResponse("")
    response["HX-Trigger"] = json.dumps(triggers)
    return response

@login_required
@user_passes_test(is_admin)
def ticket_sign(request, ticket_id):
    """
    Ký xác nhận để chuyển bước.
    Trong bản demo, Admin có thể ký thay tất cả các bước.
    """
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid method")
        
    from .models import CorrectionTicket, TicketSignature
    ticket = get_object_or_404(CorrectionTicket, pk=ticket_id)
    
    status_flow = ['DISPATCHER', 'STATION', 'TECH', 'SUPERVISOR', 'ADMIN', 'RESOLVED']
    current_idx = status_flow.index(ticket.status)
    
    if current_idx >= len(status_flow) - 1:
        return HttpResponseBadRequest("Ticket already resolved")
        
    next_status = status_flow[current_idx + 1]
    
    # Lưu chữ ký giả lập
    import hashlib
    from django.utils import timezone
    raw_sig = f"{request.user.username}-{ticket.status}-{timezone.now().timestamp()}"
    sig_hash = hashlib.sha256(raw_sig.encode()).hexdigest()
    
    TicketSignature.objects.create(
        ticket=ticket,
        signer_name=request.user.get_full_name() or request.user.username,
        role=ticket.status, # Role tương ứng với bước hiện tại
        notes=request.POST.get('notes', ''),
        signature_hash=sig_hash
    )
    
    ticket.status = next_status
    ticket.save()
    
    import json
    triggers = {}
    
    # Nếu đến bước cuối (RESOLVED), hoàn thành luồng và khôi phục rơ-le
    if next_status == 'RESOLVED':
        relay = ticket.relay
        relay.is_paused_for_correction = False
        
        # Khôi phục dữ liệu hẹn giờ nếu có
        if relay.paused_schedule_data:
            relay.check_interval_value = relay.paused_schedule_data.get('check_interval_value', 1)
            relay.check_interval_unit = relay.paused_schedule_data.get('check_interval_unit', 'h')
            relay.paused_schedule_data = None
            
        # Cập nhật lại next_check_at
        relay.last_checked_at = timezone.now()
        relay.next_check_at = relay.calculate_next_check(from_time=relay.last_checked_at)
        relay.save()
        
        triggers["notify"] = {"title": "Hoàn tất", "message": f"Đã khôi phục tự động đối chiếu cho {relay.relay_code}.", "level": "success"}
        triggers[f"refreshRow-{relay.id}"] = ""
        triggers["closeResultModal"] = ""
        
        response = HttpResponse("")
        response["HX-Trigger"] = json.dumps(triggers)
        return response
        
    # Nếu chưa xong, render lại modal với trạng thái mới
    return render(request, 'stations/partials/_ticket_workflow.html', {'ticket': ticket, 'relay': ticket.relay})

@login_required
def autocheck_row(request, relay_id):
    relay = get_object_or_404(Relay, pk=relay_id)
    return render(request, 'stations/partials/_relay_row.html', {'relay': relay})
